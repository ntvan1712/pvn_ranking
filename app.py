from flask import Flask, render_template, request
from openpyxl import load_workbook
from datetime import datetime

app = Flask(__name__)

COLS = {
    "van": 2,   # Cột B
    "toan": 3,  # Cột C
    "anh": 4,   # Cột D
    "tong": 5   # Cột E
}
SBD_COL = 1    # Cột A

# Chỉ hỗ trợ các trường có mã 2 số đầu như sau:
SUPPORTED_SCHOOLS = {
    "20": "Nguyễn Đức Thuận",
    "21": "Hoàng Văn Thụ",  
    "22": "Lương Thế Vinh",
    "24": "Tống Văn Trân",  
    "26": "Phạm Văn Nghị",
    "27": "Đại An"
}

# Mapping số thí sinh theo mã trường
SCHOOL_CANDIDATES = {
    "20": 442,  # Đại An
    "22": 425,  # Nguyễn Đức Thuận
    "21": 425,  # Hoàng Văn Thụ
    "24": 604,  # Tống Văn Trân
    "26": 494,  # Phạm Văn Nghị
    "27": 317,  # Lương Thế Vinh
}


@app.route("/", methods=["GET", "POST"])
def index():
    ranks = {}
    scores = {}
    sbd = ""
    error = ""
    school_name = ""
    total_candidates = 0

    if request.method == "POST":
        sbd = request.form["sbd"].strip()
        print(f"[{datetime.now()}] Tra cứu SBD: {sbd}")

        if len(sbd) < 2:
            error = "Mã số báo danh không hợp lệ."
        else:
            school_code = sbd[:2]
            if school_code not in SUPPORTED_SCHOOLS.keys():
                error = "Chỉ hỗ trợ các trường Đại An, Phạm Văn Nghị, Lương Thế Vinh, Nguyễn Đức Thuận, Hoàng Văn Thụ, Tống Văn Trân."
            else:
                school_name = SUPPORTED_SCHOOLS[school_code]
                excel_file = f"{school_code}.xlsx"
                total_candidates = SCHOOL_CANDIDATES[school_code]

                try:
                    wb = load_workbook(excel_file)
                    ws = wb.active

                    for key, col in COLS.items():
                        all_scores = []
                        for row in ws.iter_rows(min_row=2):
                            sbd_cell = str(row[SBD_COL - 1].value).strip()
                            raw_score = row[col - 1].value
                            try:
                                score = float(str(raw_score).replace(",", "."))
                                all_scores.append((sbd_cell, score))
                            except:
                                continue

                        # Sắp xếp giảm dần
                        all_scores.sort(key=lambda x: x[1], reverse=True)

                        # Tính dense ranking
                        rank_map = {}
                        last_score = None
                        last_rank = 0
                        actual_rank = 0

                        for sbd_item, score in all_scores:
                            actual_rank += 1
                            if score != last_score:
                                last_rank = actual_rank
                                last_score = score
                            rank_map[sbd_item] = last_rank

                        if sbd in rank_map:
                            ranks[key] = rank_map[sbd]
                            scores[key] = dict(all_scores)[sbd]
                except FileNotFoundError:
                    error = f"Chỉ hỗ trợ các trường Phạm Văn Nghị, Đại An, Lương Thế Vinh, Nguyễn Đức Thuận, Hoàng Văn Thụ, Tống Văn Trân."

    return render_template(
        "index.html",
        ranks=ranks,
        scores=scores,
        sbd=sbd,
        error=error,
        school_name=school_name,
        total_candidates=total_candidates
    )



if __name__ == "__main__":
    app.run(debug=True)
